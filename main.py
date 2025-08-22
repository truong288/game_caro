import os
import telegram  #ok đánh với bót mượt
import openpyxl
from datetime import datetime
from telegram import Update, InlineKeyboardMarkup, InlineKeyboardButton
from telegram.ext import (ApplicationBuilder, CommandHandler,
                          CallbackQueryHandler, ContextTypes)
from dotenv import load_dotenv
from telegram.ext import MessageHandler, filters
from stay_alive import keep_alive
import asyncio
import numpy as np
import math
import random
from telegram import User
from telegram.ext import ConversationHandler, MessageHandler, filters
import json
import logging

keep_alive()

logging.basicConfig(
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s',
    level=logging.INFO)
logger = logging.getLogger(__name__)


async def safe_send_message(context,
                            chat_id,
                            text,
                            reply_markup=None,
                            parse_mode=None):
    try:
        return await context.bot.send_message(chat_id=chat_id,
                                              text=text,
                                              reply_markup=reply_markup,
                                              parse_mode=parse_mode)
    except RetryAfter as e:
        print(f"Flood control exceeded. Retry in {e.retry_after} seconds")
        await asyncio.sleep(e.retry_after)
        return await safe_send_message(context, chat_id, text, reply_markup,
                                       parse_mode)
    except Exception as e:
        print(f"Error sending message: {e}")
        return None


BROADCAST_MESSAGE = range(1)
# ==================== GLOBAL =====================
players = {}
games = {}
win_stats = {}
game_modes = {}

ADMIN_FILE = "admins.json"
LOG_FILE = "admin_actions.log"


def normalize_group_id(group_id):
    """
    Chuẩn hóa Group ID từ các định dạng khác nhau
    """
    if isinstance(group_id, (int, float)):
        # Xử lý số âm lớn (định dạng khoa học)
        if abs(group_id) > 1e10:
            return str(int(group_id))
        return str(group_id)
    elif isinstance(group_id, str):
        # Xử lý chuỗi khoa học
        if 'E+' in group_id or 'e+' in group_id:
            try:
                return str(int(float(group_id)))
            except:
                return group_id
        return group_id
    else:
        return str(group_id)


def load_admins():
    try:
        with open(ADMIN_FILE, "r") as f:
            return json.load(f)
    except:
        return []


def save_admins(admins):
    with open(ADMIN_FILE, "w") as f:
        json.dump(admins, f)


def log_action(action, actor_id, target_id):
    os.makedirs("logs", exist_ok=True)
    with open(LOG_FILE, "a") as f:
        timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        f.write(
            f"[{timestamp}] {action} - By: {actor_id}, Target: {target_id}\n")


async def add_admin(update: Update, context: ContextTypes.DEFAULT_TYPE):
    admins = load_admins()
    user = update.effective_user

    if user.id not in admins:
        await update.message.reply_text("⛔ Bạn không phải Admin.")
        return

    if not context.args:
        await update.message.reply_text(
            "⚠️ Gõ đúng cú pháp: /addadmin <user_id>")
        return

    try:
        new_id = int(context.args[0])
        if new_id in admins:
            await update.message.reply_text("ℹ️ ID này đã là admin.")
        else:
            admins.append(new_id)
            save_admins(admins)
            log_action("ADD_ADMIN", user.id, new_id)
            await update.message.reply_text(f"✅ Đã thêm admin: {new_id}")
    except:
        await update.message.reply_text("❌ Lỗi: Không thể thêm ID.")


async def remove_admin(update: Update, context: ContextTypes.DEFAULT_TYPE):
    admins = load_admins()
    user = update.effective_user

    if user.id not in admins:
        await update.message.reply_text("⛔ Bạn không phải Admin.")
        return

    if not context.args:
        await update.message.reply_text(
            "⚠️ Gõ đúng cú pháp: /removeadmin <user_id>")
        return

    try:
        remove_id = int(context.args[0])

        if remove_id == user.id:
            await update.message.reply_text("❌ Không thể tự xóa chính mình.")
            return

        if remove_id in admins:
            admins.remove(remove_id)
            save_admins(admins)
            log_action("REMOVE_ADMIN", user.id, remove_id)
            await update.message.reply_text(f"🗑️ Đã xóa admin: {remove_id}")
        else:
            await update.message.reply_text("⚠️ ID này không phải admin.")
    except:
        await update.message.reply_text("❌ Lỗi: Không thể xóa ID.")


async def show_my_id(update: Update, context: ContextTypes.DEFAULT_TYPE):
    user = update.effective_user
    await update.message.reply_text(f"🆔 ID của bạn là: `{user.id}`",
                                    parse_mode="Markdown")


async def admin_list(update: Update, context: ContextTypes.DEFAULT_TYPE):
    user = update.effective_user
    admins = load_admins()
    if user.id not in admins:
        await update.message.reply_text("⛔ Bạn không có quyền.")
        return
    if not admins:
        await update.message.reply_text("⚠️ Chưa có admin nào.")
        return
    msg = "📋 Danh sách Admin hiện tại:\n"
    for uid in admins:
        try:
            member = await context.bot.get_chat_member(
                update.effective_chat.id, uid)
            name = member.user.full_name
            msg += f"👤 {name} - `{uid}`\n"
        except:
            msg += f"❓ Không rõ tên - `{uid}`\n"
    await update.message.reply_text(msg, parse_mode="Markdown")


ADMIN_IDS = load_admins()


def get_possible_moves(board_np):
    moves = set()
    for y in range(board_np.shape[0]):
        for x in range(board_np.shape[1]):
            if board_np[y][x] != "▫️":  # Nếu ô không trống
                for dy in range(-1, 2):
                    for dx in range(-1, 2):
                        ny, nx = y + dy, x + dx
                        if 0 <= ny < board_np.shape[
                                0] and 0 <= nx < board_np.shape[1]:
                            if board_np[ny][nx] == "▫️":
                                moves.add((nx, ny))
    return list(moves)


def score_line_improved(line, symbol, opp):
    score = 0
    line_str = ''.join(line)
    if f"{symbol*4}" in line_str: score += 100000  # Chiến thắng
    elif f"{symbol*3}▫️" in line_str or f"▫️{symbol*3}" in line_str:
        score += 10000
    elif f"{symbol*2}▫️{symbol}" in line_str or f"{symbol}▫️{symbol*2}" in line_str:
        score += 8000
    elif f"{symbol*2}▫️▫️" in line_str or f"▫️▫️{symbol*2}" in line_str or f"▫️{symbol*2}▫️" in line_str:
        score += 3000
    # Điểm phòng thủ (chặn đối thủ)
    if f"{opp*3}▫️" in line_str or f"▫️{opp*3}" in line_str:
        score -= 9000  # Ưu tiên chặn cao hơn
    elif f"{opp*2}▫️{opp}" in line_str or f"{opp}▫️{opp*2}" in line_str:
        score -= 8500
    elif f"▫️{opp*2}▫️" in line_str:
        score -= 4000
    elif f"{opp*2}▫️▫️" in line_str or f"▫️▫️{opp*2}" in line_str:
        score -= 2000

    return score


# Hàm đánh giá tổng thể bàn cờ
def evaluate_board(board_np, symbol):
    score = 0
    opp = "⭕" if symbol == "❌" else "❌"
    # 1. Tăng điểm cho các chuỗi dài
    for row in board_np:
        score += score_line_improved(row, symbol, opp)
    for col in board_np.T:
        score += score_line_improved(col, symbol, opp)
    for i in range(-board_np.shape[0] + 1, board_np.shape[1]):
        score += score_line_improved(np.diag(board_np, k=i), symbol, opp)
        score += score_line_improved(np.diag(np.fliplr(board_np), k=i), symbol,
                                     opp)  # Sửa ở đây
    # 2. Ưu tiên vị trí trung tâm (quan trọng hơn)
    center_positions = [(3, 4), (4, 3), (3, 3), (4, 4), (2, 2), (5, 5), (2, 5),
                        (5, 2)]
    for x, y in center_positions:
        if 0 <= y < len(board_np) and 0 <= x < len(board_np[0]):
            if board_np[y][x] == symbol:
                score += 500  # Tăng điểm lên gần gấp đôi
            elif board_np[y][x] == opp:
                score -= 400  # Trừ điểm đối thủ chiếm vị trí tốt
    # 3. Thêm điểm cho các ô xung quanh đã có quân
    for y in range(len(board_np)):
        for x in range(len(board_np[0])):
            if board_np[y][x] == symbol:
                for dy in [-1, 0, 1]:
                    for dx in [-1, 0, 1]:
                        nx, ny = x + dx, y + dy
                        if 0 <= ny < len(board_np) and 0 <= nx < len(
                                board_np[0]):
                            if board_np[ny][nx] == "▫️":
                                score += 50  # Ưu tiên đánh gần quân mình
    return score


# Thuật toán Minimax với Alpha-Beta Pruning
def minimax(board_np, depth, alpha, beta, is_maximizing, symbol, opp):
    if depth == 0:
        score = evaluate_board(board_np, symbol)
        return score, None
    moves = get_possible_moves(board_np)
    if not moves:
        return 0, None

    # Sắp xếp các nước đi theo độ ưu tiên để alpha-beta pruning hiệu quả hơn
    moves = order_moves(board_np, moves, symbol, opp)
    best_move = None
    if is_maximizing:
        max_eval = -math.inf
        for x, y in moves:
            board_np[y][x] = symbol
            eval, _ = minimax(board_np, depth - 1, alpha, beta, False, symbol,
                              opp)
            board_np[y][x] = "▫️"

            if eval > max_eval:
                max_eval = eval
                best_move = (x, y)

            alpha = max(alpha, eval)
            if beta <= alpha:
                break

        return max_eval, best_move
    else:
        min_eval = math.inf
        for x, y in moves:
            board_np[y][x] = opp
            eval, _ = minimax(board_np, depth - 1, alpha, beta, True, symbol,
                              opp)
            board_np[y][x] = "▫️"

            if eval < min_eval:
                min_eval = eval
                best_move = (x, y)

            beta = min(beta, eval)
            if beta <= alpha:
                break

        return min_eval, best_move


def order_moves(board_np, moves, symbol, opp):
    """Sắp xếp các nước đi theo độ ưu tiên"""
    scored_moves = []
    for x, y in moves:
        score = 0
        # Ưu tiên các nước đi gần quân đã có
        for dy in range(-2, 3):
            for dx in range(-2, 3):
                nx, ny = x + dx, y + dy
                if (0 <= ny < board_np.shape[0] and 0 <= nx < board_np.shape[1]
                        and board_np[ny][nx] != "▫️"):
                    score += 10
        # Ưu tiên trung tâm
        center_x, center_y = board_np.shape[1] // 2, board_np.shape[0] // 2
        distance_to_center = math.sqrt((x - center_x)**2 + (y - center_y)**2)
        score += int(20 - distance_to_center)

        scored_moves.append((score, (x, y)))
    # Sắp xếp theo điểm số giảm dần
    scored_moves.sort(key=lambda x: x[0], reverse=True)
    return [move for score, move in scored_moves]


def best_move(board,
              symbol,
              win_condition=4,
              depth=3):  # Giảm depth vì bàn cờ lớn
    board_np = np.array(board)
    opp = "⭕" if symbol == "❌" else "❌"

    # Ưu tiên 1: Kiểm tra chiến thắng ngay lập tức
    for x, y in get_possible_moves(board_np):
        board_np[y][x] = symbol
        if check_win(board_np.tolist(), symbol, win_condition):
            board_np[y][x] = "▫️"
            return (x, y)
        board_np[y][x] = "▫️"
    # Ưu tiên 2: Chặn đối thủ chiến thắng
    for x, y in get_possible_moves(board_np):
        board_np[y][x] = opp
        if check_win(board_np.tolist(), opp, win_condition):
            board_np[y][x] = "▫️"
            return (x, y)
        board_np[y][x] = "▫️"
    # Ưu tiên 3: Chiến lược tấn công - tạo các mối đe dọa đôi
    attack_moves = []
    for x, y in get_possible_moves(board_np):
        board_np[y][x] = symbol
        # Đếm số lượng mối đe dọa (các dãy 3 quân có thể thành 4)
        threats = count_threats(board_np, symbol, win_condition)
        board_np[y][x] = "▫️"
        if threats >= 2:  # Nếu tạo được 2 mối đe dọa trở lên
            attack_moves.append(((x, y), threats))
    if attack_moves:
        # Chọn nước đi tạo nhiều mối đe dọa nhất
        attack_moves.sort(key=lambda x: x[1], reverse=True)
        return attack_moves[0][0]
    # Ưu tiên 4: Chiến lược phòng thủ - chặn các mối đe dọa của đối thủ
    defense_moves = []
    for x, y in get_possible_moves(board_np):
        board_np[y][x] = opp
        threats = count_threats(board_np, opp, win_condition)
        board_np[y][x] = "▫️"
        if threats >= 1:
            defense_moves.append(((x, y), threats))
    if defense_moves:
        defense_moves.sort(key=lambda x: x[1], reverse=True)
        return defense_moves[0][0]
    # Ưu tiên 5: Chiếm vị trí trung tâm và các vị trí chiến lược
    center_positions = [(3, 4), (4, 3), (3, 3), (4, 4), (4, 5), (5, 4), (2, 3),
                        (3, 2), (5, 3), (3, 5), (2, 4), (4, 2), (5, 5), (2, 2),
                        (5, 2), (2, 5)]

    for pos in center_positions:
        x, y = pos
        if (x, y) in get_possible_moves(board_np) and board_np[y][x] == "▫️":
            return (x, y)
    # Ưu tiên 6: Sử dụng Minimax cho các nước đi phức tạp
    _, move = minimax(board_np, depth, -math.inf, math.inf, True, symbol, opp)
    if move:
        return move
    # Cuối cùng: Chọn ngẫu nhiên nếu không có nước đi tốt
    return random.choice(get_possible_moves(board_np))


def count_threats(board_np, symbol, win_condition):
    """Đếm số lượng mối đe dọa (các dãy có thể chiến thắng)"""
    threats = 0
    size_y, size_x = board_np.shape
    # Kiểm tra các hướng
    directions = [(1, 0), (0, 1), (1, 1), (1, -1)]

    for y in range(size_y):
        for x in range(size_x):
            for dx, dy in directions:
                # Kiểm tra xem có thể tạo thành dãy win_condition quân không
                if can_make_line(board_np, x, y, dx, dy, symbol,
                                 win_condition):
                    threats += 1
    return threats


def can_make_line(board_np, start_x, start_y, dx, dy, symbol, win_condition):
    """Kiểm tra xem có thể tạo thành dãy chiến thắng từ vị trí này không"""
    count = 0
    empty_count = 0
    size_y, size_x = board_np.shape

    for i in range(win_condition):
        x = start_x + i * dx
        y = start_y + i * dy

        if not (0 <= x < size_x and 0 <= y < size_y):
            return False

        if board_np[y][x] == symbol:
            count += 1
        elif board_np[y][x] == "▫️":
            empty_count += 1
        else:
            return False

    return count == win_condition - 1 and empty_count == 1


# ============== SAVE TO EXCEL ==============
async def save_player_to_excel(name, username, user_id, group_id, time_joined):
    path = "players.xlsx"
    if not os.path.exists(path):
        wb = openpyxl.Workbook()
        sheet = wb.active
        sheet.append(
            ["Tên", "Username", "User ID", "Group ID", "Thời gian tham gia"])
        wb.save(path)

    wb = openpyxl.load_workbook(path)
    sheet = wb.active

    if isinstance(group_id, (int, float)):
        actual_group_id = str(
            int(group_id)) if abs(group_id) > 1e10 else str(group_id)
    else:
        actual_group_id = str(group_id)

    for row in sheet.iter_rows(min_row=2, values_only=True):
        if len(row) < 4:  # Đảm bảo row có đủ phần tử
            continue

        existing_user_id = str(row[2]) if row[2] is not None else ""
        existing_group_id = str(row[3]) if row[3] is not None else ""

        if existing_group_id.startswith('-') and 'E+' in existing_group_id:
            try:
                existing_group_id = str(int(float(existing_group_id)))
            except:
                pass

        if existing_user_id == str(
                user_id) and existing_group_id == actual_group_id:
            print(
                f"User ID {user_id} với Group ID {actual_group_id} đã tồn tại, không thêm nữa."
            )
            wb.close()
            return

    sheet.append([
        name,
        username,
        user_id,
        actual_group_id,  # Lưu Group ID đã được xử lý
        time_joined.strftime("%d-%m-%Y %H:%M:%S")
    ])

    for cell in sheet[sheet.max_row]:
        if cell.column == 4:  # Cột Group ID
            cell.number_format = '@'  # Định dạng文本

    wb.save(path)
    wb.close()


# ============== BÀN CỜ ==============
def create_board_keyboard(board):
    keyboard = []
    for y, row in enumerate(board):
        line = [
            InlineKeyboardButton(text=cell, callback_data=f"{x},{y}")
            for x, cell in enumerate(row)
        ]
        keyboard.append(line)
    return InlineKeyboardMarkup(keyboard)


async def update_board_message(context, chat_id, show_turn=True):
    game = games.get(chat_id)
    if not game:
        return

    board = game["board"]
    markup = create_board_keyboard(board)

    if show_turn:
        current_player = game["players"][game["turn"]]
        symbol = "❌" if game["turn"] == 0 else "⭕"

        if current_player == "bot":
            message = f"🤖 Đến lượt Bot ({symbol})"
        else:
            username = f"@{current_player.username}" if current_player.username else current_player.first_name
            message = f"👤 Đến lượt {username} ({symbol})"
    else:
        message = "🎯 Trận đấu kết thúc!"

    try:
        await context.bot.edit_message_text(chat_id=chat_id,
                                            message_id=game["message_id"],
                                            text=message,
                                            reply_markup=markup)
    except:
        pass

    if game.get("task"):
        game["task"].cancel()

    if show_turn:
        game["task"] = asyncio.create_task(turn_timeout(context, chat_id))


# ============== TIMEOUT ==============
async def turn_timeout(context, chat_id):
    await asyncio.sleep(60)

    game = games.get(chat_id)
    if not game or len(game["players"]) < 2:
        return
    if game.get("task"):
        game["task"].cancel()

    loser_index = game["turn"]
    winner_index = 1 - loser_index
    loser = game["players"][loser_index]
    winner = game["players"][winner_index]
    loser_name = "Bot" if loser == "bot" else loser.full_name
    winner_name = "Bot" if winner == "bot" else winner.full_name
    winner_id = 0 if winner == "bot" else winner.id
    if winner_id not in win_stats:
        win_stats[winner_id] = {"name": winner_name, "count": 1}
    else:
        win_stats[winner_id]["count"] += 1
    markup = create_board_keyboard(game["board"])
    try:
        await context.bot.edit_message_text(
            chat_id=chat_id,
            message_id=game["message_id"],
            text="🎯 Trận đấu kết thúc do hết giờ!",
            reply_markup=markup)
    except:
        pass
    win_condition = game.get("win_condition", 4)
    winner_symbol = "❌" if winner_index == 0 else "⭕"

    win_message = (f"⏱ {loser_name} Hết thời gian!\n"
                   f"🏆 CHIẾN THẮNG! 🏆\n"
                   f"👑 {winner_name} ({winner_symbol})\n"
                   f"📌 Chế độ: {win_condition} nước thắng\n"
                   f"📊 Thắng: {win_stats[winner_id]['count']}")

    await context.bot.send_message(chat_id=chat_id, text=win_message)
    games.pop(chat_id, None)
    players.pop(chat_id, None)
    await context.bot.send_message(chat_id=chat_id,
                                   text="👉 Gõ /startgame Bắt đầu ván mới...")


# ============== COMMAND HANDLERS ==============
async def start_game(update: Update, context: ContextTypes.DEFAULT_TYPE):
    chat_id = update.effective_chat.id
    user = update.effective_user
    if chat_id in games:
        game = games[chat_id]
        if (len(game.get("players", [])) == 1 and "message_id" not in game
                and game["players"][0].id == user.id):
            games.pop(chat_id, None)
            players.pop(chat_id, None)
        elif (len(game.get("players", [])) == 1 and "message_id" not in game
              and game["players"][0].id != user.id):

            await context.bot.send_message(
                chat_id=chat_id,
                text=
                "⚠️ Đang chờ người chơi thứ 2 tham gia. Bạn có thể dùng /join")
            return
        elif len(game.get("players", [])) >= 2:
            # Kiểm tra nếu người dùng đã tham gia game này
            for player in game.get("players", []):
                if hasattr(player, 'id') and player.id == user.id:
                    await context.bot.send_message(
                        chat_id=chat_id, text="⚠️ Bạn đang tham gia.")
                    return

            await context.bot.send_message(
                chat_id=chat_id,
                text="⚠️ Phòng này đang chơi, vui lòng chờ kết thúc.")
            return
    await save_player_to_excel(user.full_name, user.username, user.id, chat_id,
                               datetime.now())
    keyboard = InlineKeyboardMarkup([[
        InlineKeyboardButton("Tham gia 4 nước thắng", callback_data="join_4")
    ], [InlineKeyboardButton("Tham gia 5 nước thắng", callback_data="join_5")],
                                     [
                                         InlineKeyboardButton(
                                             "Chơi với bot (4 nước thắng)",
                                             callback_data="join_bot")
                                     ]])

    await context.bot.send_message(chat_id=chat_id,
                                   text="🎮 Chọn chế độ chơi:",
                                   reply_markup=keyboard)


def check_game_ended(game):
    """Kiểm tra xem game đã kết thúc chưa"""
    board = game.get("board", [])
    win_condition = game.get("win_condition", 4)

    # Kiểm tra cả hai symbol
    if check_win(board, "❌", win_condition) or check_win(
            board, "⭕", win_condition):
        return True
    return False


async def join_game(update: Update, context: ContextTypes.DEFAULT_TYPE):
    chat_id = update.effective_chat.id
    user = update.effective_user

    if chat_id not in games:
        await update.message.reply_text("👉 Gõ /startgame để bắt đầu ván mới.")
        return

    game = games[chat_id]
    for player in game["players"]:
        if hasattr(player, 'id') and player.id == user.id:
            await update.message.reply_text("⛔ Bạn đã tham gia rồi!")
            return
    if "message_id" in game:
        await update.message.reply_text("⚠️ Phòng đã đủ 2 người chơi!")
        return
    if len(game["players"]) >= 2:
        await update.message.reply_text("⚠️ Phòng đã đủ 2 người chơi!")
        return
    game["players"].append(user)
    players.setdefault(chat_id, []).append(user.id)

    await save_player_to_excel(user.full_name, user.username, user.id, chat_id,
                               datetime.now())
    if len(game["players"]) == 2:
        player1 = game["players"][0]
        player2 = game["players"][1]
        win_condition = game_modes.get(chat_id, 4)

        text = f"{player1.first_name} ❌ vs {player2.first_name} ⭕\n"
        text += f"Chế độ {win_condition} nước thắng\n"
        text += f"Đến lượt: {player1.first_name} ❌"

        msg = await context.bot.send_message(
            chat_id=chat_id,
            text=text,
            reply_markup=create_board_keyboard(game["board"]))

        game["message_id"] = msg.message_id
        game["task"] = asyncio.create_task(turn_timeout(context, chat_id))


async def show_game_board(context, chat_id, update=None):
    game = games.get(chat_id)
    if not game:
        return

    if 'turn' not in game:
        game['turn'] = 0
    win_condition = game["win_condition"]

    player1 = game["players"][0]
    player2 = game["players"][1] if len(game["players"]) > 1 else "bot"

    text = f"🎮 Chế độ {win_condition} nước thắng\n"
    text += f"👤 {player1.first_name} ❌ vs "
    text += f"👤 {player2.first_name} ⭕" if player2 != "bot" else "🤖 Bot ⭕"

    if update and len(game["players"]) == 1:
        await update.message.reply_text(f"✅ {player1.first_name} đã tham gia!")

    if "message_id" in game:
        try:
            await context.bot.edit_message_text(
                chat_id=chat_id,
                message_id=game["message_id"],
                text=text,
                reply_markup=create_board_keyboard(game["board"]))
        except:
            msg = await context.bot.send_message(
                chat_id=chat_id,
                text=text,
                reply_markup=create_board_keyboard(game["board"]))
            game["message_id"] = msg.message_id
    else:
        msg = await context.bot.send_message(
            chat_id=chat_id,
            text=text,
            reply_markup=create_board_keyboard(game["board"]))
        game["message_id"] = msg.message_id

    # Thiết lập timeout
    if game.get("task"):
        game["task"].cancel()
    game["task"] = asyncio.create_task(turn_timeout(context, chat_id))


async def join_bot(update: Update, context: ContextTypes.DEFAULT_TYPE):
    chat_id = update.effective_chat.id
    user = update.effective_user

    if chat_id not in games:
        await context.bot.send_message(chat_id=chat_id,
                                       text="👉 Gõ /startgame Bắt đầu ván mới.")
        return

    players.setdefault(chat_id, [])

    if user.id not in players[chat_id]:
        if "bot" in games[chat_id]["players"]:
            await update.message.reply_text("🤖 Bạn đang chơi với bot!")
            return

        players[chat_id].append(user.id)
        games[chat_id]["players"].append(user)
        games[chat_id]["players"].append("bot")
        games[chat_id]["bot_play"] = True

        await context.bot.send_message(
            chat_id=chat_id, text=f"✅ {user.first_name} chơi với bot!")

        await save_player_to_excel(user.full_name, user.username, user.id,
                                   chat_id, datetime.now())
        current_player = games[chat_id]["players"][0]
        username = f"@{current_player.username}" if current_player != "bot" and current_player.username else current_player.first_name

        msg = await context.bot.send_message(
            chat_id=chat_id,
            text=f"Đến lượt: {username}",
            reply_markup=create_board_keyboard(games[chat_id]["board"]))

        games[chat_id]["message_id"] = msg.message_id
        games[chat_id]["task"] = asyncio.create_task(
            turn_timeout(context, chat_id))


async def reset_game(update: Update, context: ContextTypes.DEFAULT_TYPE):
    chat_id = update.effective_chat.id
    user_id = update.effective_user.id
    path = "players.xlsx"

    if user_id in ADMIN_IDS:
        try:
            win_stats.clear()
            await context.bot.send_message(chat_id=chat_id,
                                           text="🛑 Admin đã reset.")
        except Exception as e:
            await context.bot.send_message(chat_id=chat_id,
                                           text=f"❌ Lỗi khi reset: {str(e)}")
        return

    try:
        if chat_id in games:
            if games[chat_id].get("task"):
                games[chat_id]["task"].cancel()
            games.pop(chat_id, None)
            players.pop(chat_id, None)

        to_delete = []
        for uid in win_stats:
            try:
                member = await context.bot.get_chat_member(chat_id, uid)
                if member.status in ("member", "administrator", "creator"):
                    to_delete.append(uid)
            except:
                continue
        for uid in to_delete:
            win_stats.pop(uid, None)

        await context.bot.send_message(chat_id=chat_id,
                                       text="♻️ Đã reset game và thống kê!")
    except Exception as e:
        await context.bot.send_message(chat_id=chat_id,
                                       text=f"❌ Lỗi khi reset: {str(e)}")


async def export_data(update: Update, context: ContextTypes.DEFAULT_TYPE):
    chat_id = update.effective_chat.id
    path = "data/players.xlsx"

    try:
        if not os.path.exists(path):
            await context.bot.send_message(chat_id=chat_id,
                                           text="⚠️ Chưa có dữ liệu.")
            return

        with open(path, "rb") as file:
            await context.bot.send_document(chat_id=chat_id,
                                            document=file,
                                            filename="players_data.xlsx",
                                            caption="📊 Dữ liệu người chơi")
    except Exception as e:
        await context.bot.send_message(
            chat_id=chat_id, text=f"❌ Lỗi khi xuất dữ liệu: {str(e)}")


async def delete_export(update: Update, context: ContextTypes.DEFAULT_TYPE):
    chat_id = update.effective_chat.id
    path = "data/players.xlsx"

    try:
        if not os.path.exists(path):
            await context.bot.send_message(chat_id=chat_id,
                                           text="⚠️ Không tìm thấy file.")
            return

        os.remove(path)
        await context.bot.send_message(chat_id=chat_id,
                                       text="🗑️ Đã xóa file thành công!")
    except Exception as e:
        await context.bot.send_message(chat_id=chat_id,
                                       text=f"❌ Lỗi khi xóa file: {str(e)}")


async def show_win_stats(update: Update, context: ContextTypes.DEFAULT_TYPE):
    user = update.effective_user
    chat_id = update.effective_chat.id

    admins = load_admins()
    if user.id in admins:
        if not win_stats:
            await context.bot.send_message(chat_id=chat_id,
                                           text="📊 Chưa có ai thắng.")
            return

        msg = f"🏆 THỐNG KÊ TOÀN BỘ\n📍Group ID: {normalize_group_id(chat_id)}\n\n"
        for uid, data in win_stats.items():
            name = data["name"]
            count = data["count"]
            msg += f"👤 {name}: {count} Ván\n"
        await context.bot.send_message(chat_id=chat_id, text=msg)
    else:
        result = {}
        for uid, data in win_stats.items():
            try:
                member = await context.bot.get_chat_member(chat_id, uid)
                if member.status in ("member", "administrator", "creator"):
                    result[uid] = data
            except:
                pass

        if not result:
            await context.bot.send_message(chat_id=chat_id,
                                           text="📊 Nhóm chưa có ai thắng.")
            return

        msg = f"🏆 BẢNG XẾP HẠNG 🏆:\n"
        for uid, data in result.items():
            name = data["name"]
            count = data["count"]
            msg += f"{name}: {count} Ván\n"
        await context.bot.send_message(chat_id=chat_id, text=msg)


async def show_rules(update: Update, context: ContextTypes.DEFAULT_TYPE):
    chat_id = update.effective_chat.id
    await context.bot.send_message(
        chat_id=chat_id,
        text="📘 Hướng dẫn:\n\n"
        "🔹/startgame - Bắt đầu game mới.\n"
        "🔹/join - Tham gia game.\n"
        "🔹/joinbot - Tham gia chơi với bot.\n"
        "🔹/reset - Làm mới game nhóm này.\n"
        "🔹/win - Xem thống kê thắng.\n"
        "🔹/help - Xem hướng dẫn.\n"
        "🔹/admin - Admin.\n\n"
        "📌 LUẬT CHƠI:\n\n"
        "- Khi người 1 chọn chế độ tham gia, người thứ 2 chỉ cần ấn join bàn tự hiện lên.\n"
        "- Tùy vào chế độ 4 hoặc 5 điểm thẳng hàng giành chiến .\n"
        "- Chế độ đánh với bót mặc đinh 4 điểm thắng.\n"
        "👉 @xukaxuka2k1 codefree,fastandsecure👈")


async def admin_commands(update: Update, context: ContextTypes.DEFAULT_TYPE):
    user_id = update.effective_user.id
    admins = load_admins()

    if user_id not in admins:
        await update.message.reply_text("⛔ Bạn không có quyền.")
        return

    text = ("🛠️ **LỆNH ADMIN** 🛠️\n\n"
            "🔹 /myid - Lấy ID.\n"
            "🔹 /fast - Xuất file.\n"
            "🔹 /secure - Xóa file.\n"
            "🔹 /addadmin <id> - Thêm admin.\n"
            "🔹 /removeadmin <id> - Xóa admin.\n"
            "🔹 /adminlist - Danh sách admin.\n"
            "🔹 /broadcast - Gửi tin nhắn đến all.")

    await update.message.reply_text(text, parse_mode="Markdown")


# ============== XỬ LÝ NÚT NHẤN ==============
async def handle_mode_selection(update: Update,
                                context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    await query.answer()
    chat_id = query.message.chat.id
    user = query.from_user
    try:
        # KIỂM TRA TỒN TẠI TRƯỚC KHI TRUY CẬP
        if (chat_id in games and "players" in games[chat_id] and any(
                isinstance(p, User) and p.id == user.id
                for p in games[chat_id]["players"])):
            await query.edit_message_text("⛔ Bạn đã tham gia phòng này rồi!")
            return

        mode = query.data.split("_")[1]
        win_condition = 4 if mode == "bot" else int(mode)
        game_modes[chat_id] = win_condition

        if mode == "bot":
            games[chat_id] = {
                "board": [["▫️"] * 8 for _ in range(10)],
                "players": [user, "bot"],
                "turn": 0,
                "win_condition": win_condition,
                "bot_play": True,
                "created_at": datetime.now()
            }
            players[chat_id] = [user.id]

            await save_player_to_excel(user.full_name, user.username, user.id,
                                       chat_id, datetime.now())

            await asyncio.sleep(1)
            await query.delete_message()
            await asyncio.sleep(1)

            current_player = games[chat_id]["players"][0]
            username = current_player.first_name
            msg = await context.bot.send_message(
                chat_id=chat_id,
                text=f"🎮 Chế độ {win_condition} nước thắng\n"
                f"👤 {username} ❌ vs 🤖 Bot ⭕\n"
                f"Đến lượt: {username} ❌",
                reply_markup=create_board_keyboard(games[chat_id]["board"]))

            games[chat_id]["message_id"] = msg.message_id
            games[chat_id]["task"] = asyncio.create_task(
                turn_timeout(context, chat_id))

        else:
            if (chat_id in games
                    and len(games[chat_id].get("players", [])) == 1
                    and "message_id" not in games[chat_id]):
                # Xóa game cũ nếu đang chờ người thứ 2
                games.pop(chat_id, None)
                players.pop(chat_id, None)

            games[chat_id] = {
                "board": [["▫️"] * 8 for _ in range(10)],
                "players": [user],
                "turn": 0,
                "win_condition": win_condition,
                "bot_play": False,
                "created_at": datetime.now()
            }
            players[chat_id] = [user.id]

            await save_player_to_excel(user.full_name, user.username, user.id,
                                       chat_id, datetime.now())

            await query.edit_message_text(
                text=f"✅ {user.first_name} Tham gia {mode} nước thắng!\n"
                "👉 Mời người thứ 2 gõ /join\n"
                "🔄 Nếu muốn đổi chế độ, gõ /startgame lại")

    except telegram.error.RetryAfter as e:
        print(f"Flood control exceeded. Retry in {e.retry_after} seconds")
        await asyncio.sleep(e.retry_after)
        await handle_mode_selection(update, context)
    except Exception as e:
        print(f"Error in handle_mode_selection: {e}")
        await query.edit_message_text("❌ Đã xảy ra lỗi, vui lòng thử lại.")


async def handle_move(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    await query.answer()
    chat_id = query.message.chat.id
    user = query.from_user
    game = games.get(chat_id)

    if not game:
        return

    current_player = game["players"][game["turn"]]
    if current_player != user and current_player != "bot":
        await context.bot.send_message(chat_id=chat_id,
                                       text="⛔ Chưa đến lượt bạn!")
        return
    try:
        x, y = map(int, query.data.split(","))
    except ValueError:
        return
    if game["board"][y][x] != "▫️":
        await context.bot.send_message(chat_id=chat_id,
                                       text="❗ Ô này đã được đánh rồi!")
        return

    symbol = "❌" if current_player == game["players"][0] else "⭕"
    game["board"][y][x] = symbol

    if game.get("task"):
        game["task"].cancel()

    win_condition = game.get("win_condition", 4)

    # Hàm xử lý chiến thắng chung
    async def handle_win(winner, symbol_used):
        winner_id = 0 if winner == "bot" else winner.id
        winner_name = "🤖 Bot" if winner == "bot" else f"👤 {winner.full_name}"
        win_stats.setdefault(winner_id, {"name": winner_name, "count": 0})
        win_stats[winner_id]["count"] += 1
        await update_board_message(context, chat_id, show_turn=False)
        await context.bot.send_message(
            chat_id=chat_id,
            text=(f"🏆 CHIẾN THẮNG! 🏆\n"
                  f"👑 {winner_name}\n"
                  f"📌 Chế độ: {win_condition} nước thắng\n"
                  f"📊 Thắng: {win_stats[winner_id]['count']}"))
        await context.bot.send_message(chat_id=chat_id,
                                       text="👉 Gõ /startgame Bắt đầu ván mới.")
        games.pop(chat_id, None)
        players.pop(chat_id, None)

    if check_win(game["board"], symbol, win_condition):
        await handle_win(current_player, symbol)
        return

    game["turn"] = 1 - game["turn"]
    await update_board_message(context, chat_id, show_turn=True)

    if game.get("bot_play") and game["players"][game["turn"]] == "bot":
        await asyncio.sleep(1)
        thinking_msg = await context.bot.send_message(
            chat_id=chat_id, text="🤖 Bot đang suy nghĩ...")

        move = best_move(game["board"], "⭕")
        if move:
            x, y = move
            game["board"][y][x] = "⭕"
            await context.bot.delete_message(
                chat_id=chat_id, message_id=thinking_msg.message_id)

            if check_win(game["board"], "⭕", 4):
                await handle_win("bot", "⭕")
                return

            game["turn"] = 0
            await update_board_message(context, chat_id, show_turn=True)


def check_win(board, symbol, win_condition=4):
    size_y = len(board)
    size_x = len(board[0])

    for y in range(size_y):
        for x in range(size_x - win_condition + 1):
            if all(board[y][x + i] == symbol for i in range(win_condition)):
                return True

    for x in range(size_x):
        for y in range(size_y - win_condition + 1):
            if all(board[y + i][x] == symbol for i in range(win_condition)):
                return True

    for y in range(size_y - win_condition + 1):
        for x in range(size_x - win_condition + 1):
            if all(board[y + i][x + i] == symbol
                   for i in range(win_condition)):
                return True

    for y in range(size_y - win_condition + 1):
        for x in range(win_condition - 1, size_x):
            if all(board[y + i][x - i] == symbol
                   for i in range(win_condition)):
                return True

    return False


async def error_handler(update: Update,
                        context: ContextTypes.DEFAULT_TYPE) -> None:
    """Log errors and send a message if possible."""
    try:
        logger.error("Exception while handling an update:",
                     exc_info=context.error)

        if update and update.effective_message:
            await safe_send_message(context,
                                    chat_id=update.effective_chat.id,
                                    text="⚠️ Đã xảy ra lỗi, vui lòng thử lại!")
    except Exception as e:
        logger.critical(f"Error in error handler: {e}", exc_info=True)


async def unknown_command(update: Update, context: ContextTypes.DEFAULT_TYPE):
    await update.message.reply_text(
        "❓ Lệnh không hợp lệ. Gõ /help xem hướng dẫn.\n\n"
        "🎮 gameCaro: @Game_carobot\n"
        "🎮 game Nối Chữ: @noi_chu_bot\n"
        "🀄️ Dịch Tiếng Trung:  @Dichngon_ngubot")


async def start_broadcast(update, context):
    admins = load_admins()
    if update.effective_user.id not in admins:
        await update.message.reply_text("⛔ Không có quyền gửi.")
        return ConversationHandler.END

    await update.message.reply_text("✏ Nhập nội dung muốn gửi:")
    return BROADCAST_MESSAGE


async def send_broadcast(update, context):
    msg = update.message.text

    path = "data/players.xlsx"
    if not os.path.exists(path):
        await update.message.reply_text("⚠️ Chưa có dữ liệu người dùng để gửi."
                                        )
        return ConversationHandler.END

    wb = openpyxl.load_workbook(path)
    sheet = wb.active

    user_ids = []
    for row in sheet.iter_rows(min_row=2, values_only=True):
        try:
            user_id = int(row[2])  # Cột thứ 3 là User ID
            user_ids.append(user_id)
        except:
            continue
    count = 0
    for uid in user_ids:
        try:
            await context.bot.send_message(chat_id=uid, text=msg)
            count += 1
            await asyncio.sleep(0.1)  # tránh spam quá nhanh
        except Exception as e:
            print(f"Không thể gửi tới {uid}: {e}")

    await update.message.reply_text(f"✅ Đã gửi broadcast đến {count} người.")
    return ConversationHandler.END


async def cancel_broadcast(update, context):
    await update.message.reply_text("❌ Đã hủy.")
    return ConversationHandler.END


# ============== MAIN ==========================
load_dotenv()
TOKEN = os.getenv("TOKEN")
app = ApplicationBuilder().token(TOKEN).build()

app.add_handler(CommandHandler("startgame", start_game))
app.add_handler(CommandHandler("join", join_game))
app.add_handler(CommandHandler("joinbot", join_bot))
app.add_handler(CommandHandler("reset", reset_game))
app.add_handler(CommandHandler("fast", export_data))
app.add_handler(CommandHandler("secure", delete_export))
app.add_handler(CommandHandler("win", show_win_stats))
app.add_handler(CommandHandler("help", show_rules))
app.add_handler(CommandHandler("admin", admin_commands))
app.add_handler(CommandHandler("addadmin", add_admin))
app.add_handler(CommandHandler("removeadmin", remove_admin))
app.add_handler(CommandHandler("adminlist", admin_list))
app.add_handler(CommandHandler("myid", show_my_id))
app.add_handler(CallbackQueryHandler(handle_mode_selection, pattern="^join_"))
app.add_handler(CallbackQueryHandler(handle_move))

app.add_handler(
    ConversationHandler(
        entry_points=[CommandHandler("broadcast", start_broadcast)],
        states={
            BROADCAST_MESSAGE:
            [MessageHandler(filters.TEXT & ~filters.COMMAND, send_broadcast)]
        },
        fallbacks=[CommandHandler("cancel", cancel_broadcast)],
    ))
app.add_handler(MessageHandler(filters.COMMAND, unknown_command))
app.add_error_handler(error_handler)

app.run_polling()
